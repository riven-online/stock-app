import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import io
import plotly.express as px

# ضبط إعدادات الصفحة
st.set_page_config(
    page_title="RIVEN AI - Futuristic Plan & Stock Engine",
    page_icon="⚡",
    layout="wide"
)

# --- تصميم واجهة مستقبلي مع تأثيرات النيون ---
st.markdown("""
<style>
    .stApp {
        background-color: #0a0e17;
        color: #e0e6ed;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    .neon-title {
        font-size: 2.8rem;
        font-weight: 800;
        text-align: center;
        color: #00f3ff;
        text-shadow: 0 0 10px #00f3ff, 0 0 20px #00f3ff, 0 0 40px #0077ff;
        margin-bottom: 5px;
        letter-spacing: 2px;
    }
    .neon-subtitle {
        text-align: center;
        color: #ff007f;
        font-size: 1.2rem;
        text-shadow: 0 0 8px #ff007f;
        margin-bottom: 30px;
    }
    div[data-testid="stMetric"] {
        background: rgba(16, 25, 44, 0.75);
        border: 1px solid #00f3ff;
        box-shadow: 0 0 15px rgba(0, 243, 255, 0.2);
        border-radius: 12px;
        padding: 15px;
        transition: all 0.3s ease-in-out;
    }
    div[data-testid="stMetric"]:hover {
        box-shadow: 0 0 25px rgba(0, 243, 255, 0.6);
        transform: translateY(-3px);
    }
    div[data-testid="stMetricLabel"] {
        color: #8b9bb4 !important;
        font-weight: 600;
    }
    div[data-testid="stMetricValue"] {
        color: #00f3ff !important;
        text-shadow: 0 0 10px rgba(0, 243, 255, 0.5);
    }
    .stButton>button {
        width: 100%;
        background: linear-gradient(135deg, #00f3ff 0%, #ff007f 100%) !important;
        color: #ffffff !important;
        font-weight: 700 !important;
        font-size: 1.1rem !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 12px 24px !important;
        box-shadow: 0 0 15px rgba(255, 0, 127, 0.4) !important;
        transition: all 0.3s ease !important;
    }
    .stButton>button:hover {
        box-shadow: 0 0 30px rgba(0, 243, 255, 0.8) !important;
        transform: scale(1.02);
    }
    .stAlert {
        background-color: rgba(0, 243, 255, 0.05);
        border: 1px solid #00f3ff;
        color: #00f3ff;
    }
</style>
""", unsafe_allow_html=True)


def process_plan_and_stock(uploaded_file):
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names
    
    plan_sheet = None
    stock_sheet = None
    
    for sheet in sheet_names:
        sheet_lower = sheet.lower()
        if 'plan' in sheet_lower or 'reallocation' in sheet_lower:
            plan_sheet = sheet
        elif 'ستوك' in sheet_lower or 'stock' in sheet_lower:
            stock_sheet = sheet

    if not plan_sheet or not stock_sheet:
        st.error("❌ تعذر التعرف التلقائي على شيت الخطة وشيت الاستوك. يرجى التأكد من المسميات.")
        return None, None, None

    df_plan = pd.read_excel(uploaded_file, sheet_name=plan_sheet)
    df_stock = pd.read_excel(uploaded_file, sheet_name=stock_sheet)

    wb = openpyxl.Workbook()
    
    # 1. ورقة الخطة
    ws_plan = wb.active
    ws_plan.title = 'Reallocation_Plan'
    ws_plan.views.sheetView[0].showGridLines = True

    # 2. ورقة الاستوك الأساسية
    ws_stock = wb.create_sheet(title='ستوك')
    ws_stock.views.sheetView[0].showGridLines = True
    ws_stock.append(['Product/Barcode', 'Quantity'])
    for row in df_stock.itertuples(index=False):
        ws_stock.append([row[0], row[1]])

    # 3. ورقة الاستوك النهائي (للترحيل للبلان القادمة)
    ws_final_stock = wb.create_sheet(title='الاستوك النهائي')
    ws_final_stock.views.sheetView[0].showGridLines = True
    ws_final_stock.append(['Product/Barcode', 'Final_Quantity'])

    # استبعاد الأعمدة الزائدة
    store_cols = [c for c in df_plan.columns[3:] if not str(c).startswith('Unnamed')]
    
    # مسميات الأعمدة الرئيسية
    new_headers = (
        ['Item-Size', 'Item', 'Size'] 
        + store_cols 
        + ['إجمالي الخطة', 'رصيد الاستوك', 'تجهيز الخطة', 'رصيد الاستوك النهائي', 'العجز', 'حالة التغطية', 'ملاحظات الدفعات']
    )
    ws_plan.append(new_headers)

    # تحديد الحروف المرجعية للأعمدة ديناميكياً
    last_store_col_idx = 3 + len(store_cols)
    col_last_store = get_column_letter(last_store_col_idx)
    
    col_total_plan = get_column_letter(last_store_col_idx + 1)
    col_stock_qty = get_column_letter(last_store_col_idx + 2)
    col_prep_qty = get_column_letter(last_store_col_idx + 3)
    col_final_stock = get_column_letter(last_store_col_idx + 4)
    col_deficit = get_column_letter(last_store_col_idx + 5)
    col_coverage = get_column_letter(last_store_col_idx + 6)
    col_batch_notes = get_column_letter(last_store_col_idx + 7)

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    for col_num in range(1, len(new_headers) + 1):
        cell = ws_plan.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    num_rows = len(df_plan)

    # قائمة الاختيارات المنسدلة (Data Validation)
    options_list = [
        "دفعة 1", "دفعة 2", "دفعة 3", "دفعة 4", "دفعة 5", 
        "دفعة 6", "دفعة 7", "دفعة 8", "دفعة 9", "دفعة 10",
        "جاري التجهيز", "مؤجل", "تم الشحن", "مكتمل", "معلق", "مطلوب استكمال", "ملغى"
    ]
    formula_options = f'"{",".join(options_list)}"'
    
    dv = DataValidation(type="list", formula1=formula_options, allow_blank=True)
    dv.error ='عفواً، اختر قيمة من القائمة المنسدلة فقط'
    dv.errorTitle = 'إدخال غير صالح'
    dv.prompt = 'اختر الدفعة أو حالة التجهيز'
    dv.promptTitle = 'ملاحظات الدفعات'
    
    ws_plan.add_data_validation(dv)
    
    # كتابة الصفوف والمعادلات
    for idx, row in enumerate(df_plan.itertuples(index=False), start=2):
        item_size = row[0]
        item = row[1]
        size = row[2]
        store_vals = list(row[3:3+len(store_cols)])
        
        total_plan_fmt = f"=SUM(D{idx}:{col_last_store}{idx})"
        stock_qty_fmt = f'=IFERROR(VLOOKUP(A{idx}, ستوك!A:B, 2, FALSE), 0)'
        
        # ترك خلية تجهيز الخطة فارغة (None) لعدم إظهار الصفر
        prepped_qty = None 
        
        # استخدام دالة N() لضمان معالجة الخلية الفارغة كـ 0 في المعادلة الحسابية
        final_stock_fmt = f'=MAX(0, {col_stock_qty}{idx}-N({col_prep_qty}{idx}))'
        deficit_fmt = f'=IF({col_total_plan}{idx}>{col_stock_qty}{idx}, {col_total_plan}{idx}-{col_stock_qty}{idx}, 0)'
        
        coverage_fmt = (
            f'=IF({col_final_stock}{idx}>{col_total_plan}{idx}, "مكتمل بالكامل + فائض مخزون", '
            f'IF({col_final_stock}{idx}={col_total_plan}{idx}, "مكتمل بالكامل", '
            f'IF({col_stock_qty}{idx}>0, "تغطية جزئية", "عجز كامل")))'
        )
        
        # ترك ملاحظات الدفعات فارغة تماماً مع بقاء القائمة المنسدلة مفعلة
        default_batch_note = None 
        
        row_data = [item_size, item, size] + store_vals + [
            total_plan_fmt, stock_qty_fmt, prepped_qty, final_stock_fmt, deficit_fmt, coverage_fmt, default_batch_note
        ]
        ws_plan.append(row_data)

        ws_final_stock.append([item_size, f"=Reallocation_Plan!{col_final_stock}{idx}"])

    # تطبيق القائمة المنسدلة على نطاق عمود ملاحظات الدفعات
    dv.add(f"{col_batch_notes}2:{col_batch_notes}{num_rows + 1}")

    # التنسيقات الشرطية (الأصفر الخفيف المخصص للاستوك المتاح)
    soft_yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    red_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    ws_plan.conditional_formatting.add(
        f"{col_stock_qty}2:{col_stock_qty}{num_rows + 1}",
        CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=False, fill=soft_yellow_fill)
    )

    ws_plan.conditional_formatting.add(
        f"{col_deficit}2:{col_deficit}{num_rows + 1}",
        CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=False, fill=red_fill)
    )

    for col in ws_plan.columns:
        max_len = max(len(str(cell.value or '')) for cell in col[:1])
        col_letter = get_column_letter(col[0].column)
        ws_plan.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, df_plan, df_stock


# --- الواجهة الرئيسية ---

st.markdown("<h1 class='neon-title'>⚡ RIVEN AI ANALYTICS ENGINE</h1>", unsafe_allow_html=True)
st.markdown("<p class='neon-subtitle'>نظام معالجة وتوليد خطط التوزيع</p>", unsafe_allow_html=True)

uploaded_file = st.file_uploader("📥 قم برفع ملف الخطة والاستوك (xlsx)", type=['xlsx'])

if uploaded_file is not None:
    if st.button("🛸 بدء التحليل البرمجي وتوليد الخطة"):
        with st.spinner("جاري قراءة البيانات، معالجة القوائم المنسدلة، وتطوير شيت الأكسيل..."):
            excel_out, df_plan, df_stock = process_plan_and_stock(uploaded_file)
            if excel_out:
                st.session_state['excel_out'] = excel_out
                st.session_state['df_plan'] = df_plan
                st.session_state['df_stock'] = df_stock
                st.success("✨ تم تحديث الملف وإخفاء الأصفار وتفريغ الملاحظات بنجاح!")

if 'excel_out' in st.session_state:
    st.divider()
    
    col_dl, col_blank = st.columns([1, 2])
    with col_dl:
        st.download_button(
            label="💾 تصدير شيت الخطة والمخزون المحدث (Excel)",
            data=st.session_state['excel_out'],
            file_name="Riven_Plan_Master_Processed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.markdown("<h2 style='color:#00f3ff; text-shadow:0 0 10px #00f3ff;'>🌐 مركز المراقبة والتحليل الشامل (Control Console)</h2>", unsafe_allow_html=True)
    
    df_plan = st.session_state['df_plan']
    df_stock = st.session_state['df_stock']
    
    stock_map = dict(zip(df_stock.iloc[:, 0], df_stock.iloc[:, 1]))
    store_cols = [c for c in df_plan.columns[3:] if not str(c).startswith('Unnamed')]
    
    df_calc = df_plan.copy()

    for c in store_cols:
        df_calc[c] = pd.to_numeric(df_calc[c], errors='coerce').fillna(0)
    
    df_calc['إجمالي الخطة'] = df_calc[store_cols].sum(axis=1)
    df_calc['رصيد الاستوك'] = pd.to_numeric(df_calc['Item-Size'].map(stock_map), errors='coerce').fillna(0)
    df_calc['تجهيز الخطة'] = 0
    df_calc['رصيد الاستوك النهائي'] = (df_calc['رصيد الاستوك'] - df_calc['تجهيز الخطة']).clip(lower=0)
    df_calc['العجز'] = (df_calc['إجمالي الخطة'] - df_calc['رصيد الاستوك']).apply(lambda x: x if x > 0 else 0)

    def assign_status(row):
        stock_fin = row['رصيد الاستوك النهائي']
        plan = row['إجمالي الخطة']
        stock_init = row['رصيد الاستوك']
        if stock_fin > plan:
            return "مكتمل بالكامل + فائض مخزون"
        elif stock_fin == plan and plan > 0:
            return "مكتمل بالكامل"
        elif stock_init > 0:
            return "تغطية جزئية"
        else:
            return "عجز كامل"

    df_calc['حالة التغطية'] = df_calc.apply(assign_status, axis=1)

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("📦 إجمالي المطلوب", f"{int(df_calc['إجمالي الخطة'].sum()):,}")
    m2.metric("🏬 الاستوك المتاح", f"{int(df_calc['رصيد الاستوك'].sum()):,}")
    m3.metric("⚙️ تم التجهيز", f"{int(df_calc['تجهيز الخطة'].sum()):,}")
    m4.metric("🔋 الاستوك النهائي", f"{int(df_calc['رصيد الاستوك النهائي'].sum()):,}")
    m5.metric("🚨 إجمالي العجز", f"{int(df_calc['العجز'].sum()):,}")

    st.markdown("---")

    c_chart1, c_chart2 = st.columns(2)
    
    with c_chart1:
        st.markdown("<h4 style='color:#ff007f;'>📊 توزيع الفروع الأكثر طلباً</h4>", unsafe_allow_html=True)
        store_sums = df_calc[store_cols].sum().sort_values(ascending=False).head(10).reset_index()
        store_sums.columns = ['الفرع', 'الكمية المطلوبة']
        fig_stores = px.bar(store_sums, x='الفرع', y='الكمية المطلوبة', color='الكمية المطلوبة',
                            color_continuous_scale='Electric', template='plotly_dark')
        st.plotly_chart(fig_stores, use_container_width=True)

    with c_chart2:
        st.markdown("<h4 style='color:#00f3ff;'>⭕ نسب تغطية الأصناف والحالات</h4>", unsafe_allow_html=True)
        status_counts = df_calc['حالة التغطية'].value_counts().reset_index()
        status_counts.columns = ['الحالة', 'العدد']
        fig_pie = px.pie(status_counts, values='العدد', names='الحالة', hole=0.4,
                         color_discrete_sequence=['#00f3ff', '#ff007f', '#0077ff', '#7928CA'], template='plotly_dark')
        st.plotly_chart(fig_pie, use_container_width=True)

    st.markdown("<h3 style='color:#00f3ff;'>🔍 الفحص الدقيق والفلترة التفاعلية</h3>", unsafe_allow_html=True)
    
    selected_status = st.selectbox(
        "تصفية الجدول حسب حالة التغطية المخزنية:",
        options=["عرض كافة الحالات"] + list(df_calc['حالة التغطية'].unique())
    )

    if selected_status != "عرض كافة الحالات":
        filtered_df = df_calc[df_calc['حالة التغطية'] == selected_status]
    else:
        filtered_df = df_calc

    display_cols = ['Item-Size', 'Item', 'Size', 'إجمالي الخطة', 'رصيد الاستوك', 'تجهيز الخطة', 'رصيد الاستوك النهائي', 'العجز', 'حالة التغطية']
    st.dataframe(filtered_df[display_cols], use_container_width=True)
